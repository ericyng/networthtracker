from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Q
from django.db import models
from django.utils import timezone
from datetime import datetime, timedelta
from django.contrib.auth.forms import UserChangeForm
from django.contrib.auth.models import User
from django.template.loader import render_to_string
import csv
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from .models import Account, Transaction, AccountEntry
from .forms import AccountForm, TransactionForm, AccountEntryForm


def landing_page(request):
    """Landing page view for non-authenticated users"""
    if request.user.is_authenticated:
        return redirect('dashboard:dashboard')
    return render(request, 'dashboard/landing.html')


def parse_month(month_value):
    """Convert month value to integer (1-12)"""
    if isinstance(month_value, int):
        return month_value
    
    month_str = str(month_value).strip().lower()
    
    # Handle numeric strings
    if month_str.isdigit():
        month_num = int(month_str)
        if 1 <= month_num <= 12:
            return month_num
    
    # Handle month names (including common misspellings)
    month_names = {
        'january': 1, 'jan': 1, 'janurary': 1,  # Handle common misspelling
        'february': 2, 'feb': 2, 'feburary': 2,  # Handle common misspelling
        'march': 3, 'mar': 3,
        'april': 4, 'apr': 4,
        'may': 5,
        'june': 6, 'jun': 6,
        'july': 7, 'jul': 7,
        'august': 8, 'aug': 8,
        'september': 9, 'sep': 9, 'sept': 9,
        'october': 10, 'oct': 10,
        'november': 11, 'nov': 11,
        'december': 12, 'dec': 12
    }
    
    return month_names.get(month_str, 1)  # Default to January if not found


def parse_balance(balance_value):
    """Convert balance value to float, handling commas and currency symbols"""
    if isinstance(balance_value, (int, float)):
        return float(balance_value)
    
    if balance_value is None or balance_value == '':
        return 0.0
    
    balance_str = str(balance_value).strip()
    
    # Remove currency symbols and commas
    balance_str = re.sub(r'[$,€£¥₹]', '', balance_str)
    balance_str = re.sub(r',', '', balance_str)
    
    # Handle negative values with parentheses (e.g., "(1,234.56)")
    if balance_str.startswith('(') and balance_str.endswith(')'):
        balance_str = '-' + balance_str[1:-1]
    
    try:
        return float(balance_str)
    except ValueError:
        print(f"Warning: Could not parse balance value '{balance_value}' as float, using 0.0")
        return 0.0


def get_month_name(month_number):
    """Convert month number to month name"""
    month_names = {
        1: 'January', 2: 'February', 3: 'March', 4: 'April',
        5: 'May', 6: 'June', 7: 'July', 8: 'August',
        9: 'September', 10: 'October', 11: 'November', 12: 'December'
    }
    return month_names.get(month_number, 'January')


@login_required
def dashboard(request):
    """Main dashboard view showing overview of accounts"""
    user = request.user
    
    # Get time period from request, default to 12 months (1 year)
    period = request.GET.get('period', '12')
    try:
        months_to_show = int(period)
        # Validate the period
        if months_to_show not in [3, 6, 12, 24, 36, 48, 60]:
            months_to_show = 12
    except ValueError:
        months_to_show = 12
    
    # Get user's accounts
    accounts = Account.objects.filter(user=user, is_active=True)
    
    # Calculate total balance across all accounts using latest entries
    total_balance = sum(account.get_latest_balance() for account in accounts)
    
    # Get time series data for the bar chart
    from django.db.models import Sum
    from datetime import datetime, timedelta
    
    # Get the current month and year
    current_date = timezone.now()
    chart_data = []
    line_chart_data = []
    
    # Generate data for the specified number of months
    for i in range(months_to_show - 1, -1, -1):  # Start from months_to_show-1 down to 0
        # Calculate the target month properly
        target_month = current_date.month - i
        target_year = current_date.year
        
        # Handle month rollover
        while target_month <= 0:
            target_month += 12
            target_year -= 1
        
        month = target_month
        year = target_year
        
        # Get total balance for this month/year (for bar chart)
        monthly_balance = AccountEntry.objects.filter(
            account__user=user,
            month=month,
            year=year
        ).aggregate(total=Sum('balance'))['total'] or 0
        
        # Get detailed breakdown for line chart
        monthly_entries = AccountEntry.objects.filter(
            account__user=user,
            month=month,
            year=year
        ).select_related('account')
        
        # Calculate category breakdown for this month
        monthly_categories = {
            'cash': 0,
            'equity_investments': 0,
            'retirement': 0,
            'property': 0,
            'debts': 0,
            'other': 0
        }
        
        for entry in monthly_entries:
            account = entry.account
            balance = entry.balance
            
            # Determine category based on account type and classification
            if account.account_type in ['checking', 'savings']:
                monthly_categories['cash'] += balance
            elif account.account_type == 'investment' and account.classification not in ['pretax', 'posttax', 'roth', 'traditional', '401k', 'hsa', '529', 'fsa']:
                monthly_categories['equity_investments'] += balance
            elif account.classification in ['pretax', 'posttax', 'roth', 'traditional', '401k', 'hsa', '529', 'fsa']:
                monthly_categories['retirement'] += balance
            elif account.asset_type == 'property':
                monthly_categories['property'] += balance
            elif account.asset_type == 'crypto':
                monthly_categories['equity_investments'] += balance
            elif account.classification == 'debts' or account.account_type in ['loan', 'credit']:
                monthly_categories['debts'] += balance
            else:
                monthly_categories['other'] += balance
        
        # Calculate net worth (all categories combined)
        monthly_net_worth = sum(monthly_categories.values())
        
        # Create proper month label
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        month_label = f"{month_names[month-1]} {year}"
        
        chart_data.append({
            'month': month_label,
            'balance': float(monthly_balance),
            'net_worth': float(monthly_net_worth),
            'cash': float(monthly_categories['cash']),
            'equity_investments': float(monthly_categories['equity_investments']),
            'retirement': float(monthly_categories['retirement']),
            'property': float(monthly_categories['property']),
            'debts': float(monthly_categories['debts']),
            'other': float(monthly_categories['other']),
            'month_num': month,
            'year': year
        })
        
        line_chart_data.append({
            'month': month_label,
            'net_worth': float(monthly_net_worth),
            'cash': float(monthly_categories['cash']),
            'equity_investments': float(monthly_categories['equity_investments']),
            'retirement': float(monthly_categories['retirement']),
            'property': float(monthly_categories['property']),
            'debts': float(monthly_categories['debts']),  # Keep as negative for proper calculation
            'other': float(monthly_categories['other']),
            'month_num': month,
            'year': year
        })
    
    # Calculate current category totals for pie chart
    current_categories = {
        'Cash': 0,
        'Equity & Investments': 0,
        'Retirement': 0,
        'Property': 0,
        'Debts': 0,
        'Other': 0
    }
    
    for account in accounts:
        balance = account.get_latest_balance()
        if account.account_type in ['checking', 'savings']:
            current_categories['Cash'] += balance
        elif account.account_type == 'investment' and account.classification not in ['pretax', 'posttax', 'roth', 'traditional', '401k', 'hsa', '529', 'fsa']:
            current_categories['Equity & Investments'] += balance
        elif account.classification in ['pretax', 'posttax', 'roth', 'traditional', '401k', 'hsa', '529', 'fsa']:
            current_categories['Retirement'] += balance
        elif account.asset_type == 'property':
            current_categories['Property'] += balance
        elif account.asset_type == 'crypto':
            current_categories['Equity & Investments'] += balance
        elif account.classification == 'debts' or account.account_type in ['loan', 'credit']:
            current_categories['Debts'] += balance
        else:
            current_categories['Other'] += balance
    
    # Filter out categories with zero values
    pie_chart_data = {k: v for k, v in current_categories.items() if v != 0}
    
    # Get period display name
    period_names = {
        3: '3 Months',
        6: '6 Months', 
        12: '1 Year',
        24: '2 Years',
        36: '3 Years',
        48: '4 Years',
        60: '5 Years'
    }
    
    context = {
        'accounts': accounts,
        'total_balance': total_balance,
        'has_accounts': accounts.exists(),
        'chart_data': chart_data,
        'line_chart_data': line_chart_data,
        'pie_chart_data': pie_chart_data,
        'current_period': months_to_show,
        'period_name': period_names.get(months_to_show, '1 Year'),
        'available_periods': [
            {'value': 3, 'name': '3 Months'},
            {'value': 6, 'name': '6 Months'},
            {'value': 12, 'name': '1 Year'},
            {'value': 24, 'name': '2 Years'},
            {'value': 36, 'name': '3 Years'},
            {'value': 48, 'name': '4 Years'},
            {'value': 60, 'name': '5 Years'},
        ]
    }
    
    return render(request, 'dashboard/dashboard.html', context)


@login_required
def accounts_list(request):
    """View to list all user accounts"""
    accounts = Account.objects.filter(user=request.user, is_active=True)
    
    # Get sorting parameters
    sort_by = request.GET.get('sort', 'name')
    order = request.GET.get('order', 'asc')
    
    # Validate sort field
    valid_sort_fields = ['name', 'account_type', 'classification', 'asset_type', 'institution', 'balance']
    if sort_by not in valid_sort_fields:
        sort_by = 'name'
    
    # Apply sorting
    if sort_by == 'balance':
        # For balance sorting, we need to annotate with the latest balance
        accounts = accounts.annotate(
            latest_balance=models.Subquery(
                AccountEntry.objects.filter(
                    account=models.OuterRef('pk')
                ).order_by('-year', '-month').values('balance')[:1]
            )
        )
        if order == 'desc':
            accounts = accounts.order_by('-latest_balance')
        else:
            accounts = accounts.order_by('latest_balance')
    else:
        if order == 'desc':
            accounts = accounts.order_by(f'-{sort_by}')
        else:
            accounts = accounts.order_by(sort_by)
    
    # Calculate summary statistics using latest entries
    total_balance = sum(account.get_latest_balance() for account in accounts)
    
    # Calculate total debts (accounts with 'debts' classification)
    total_debts = sum(
        account.get_latest_balance() 
        for account in accounts 
        if account.classification == 'debts'
    )
    
    # Calculate total assets (excluding debts and loans)
    total_assets = sum(
        account.get_latest_balance() 
        for account in accounts 
        if account.classification != 'debts' and account.account_type != 'loan'
    )
    
    # Group accounts by classification (same as Add/Edit Entry view)
    accounts_by_classification = {}
    classification_totals = {}
    
    for account in accounts:
        balance = account.get_latest_balance()
        classification = account.get_classification_display()
        
        if classification not in accounts_by_classification:
            accounts_by_classification[classification] = []
            classification_totals[classification] = 0
        accounts_by_classification[classification].append(account)
        classification_totals[classification] += balance
    
    # Debug information
    print(f"DEBUG: Found {accounts.count()} accounts for user {request.user}")
    print(f"DEBUG: Accounts by classification: {accounts_by_classification}")
    print(f"DEBUG: Classification totals: {classification_totals}")
    
    context = {
        'accounts': accounts,
        'accounts_by_classification': accounts_by_classification,
        'classification_totals': classification_totals,
        'total_balance': total_balance,
        'total_debts': total_debts,
        'total_assets': total_assets,
        'sort_by': sort_by,
        'order': order,
    }
    return render(request, 'dashboard/accounts_list.html', context)


@login_required
def account_detail(request, account_id):
    """View to show account details and transactions"""
    account = get_object_or_404(Account, id=account_id, user=request.user)
    transactions = Transaction.objects.filter(account=account).order_by('-date', '-created_at')
    
    context = {
        'account': account,
        'transactions': transactions,
    }
    return render(request, 'dashboard/account_detail.html', context)


@login_required
def account_create(request):
    """View to create a new account"""
    if request.method == 'POST':
        form = AccountForm(request.POST)
        if form.is_valid():
            account = form.save(commit=False)
            account.user = request.user
            account.save()
            messages.success(request, 'Account created successfully!')
            return redirect('dashboard:accounts_list')
    else:
        form = AccountForm()
    
    return render(request, 'dashboard/account_form.html', {'form': form, 'title': 'Add New Account'})


@login_required
def account_edit(request, account_id):
    """View to edit an existing account"""
    account = get_object_or_404(Account, id=account_id, user=request.user)
    
    if request.method == 'POST':
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            messages.success(request, 'Account updated successfully!')
            return redirect('dashboard:accounts_list')
    else:
        form = AccountForm(instance=account)
    
    return render(request, 'dashboard/account_form.html', {
        'form': form, 
        'title': 'Edit Account',
        'account': account
    })


@login_required
def account_delete(request, account_id):
    """View to delete an account"""
    account = get_object_or_404(Account, id=account_id, user=request.user)
    
    if request.method == 'POST':
        account.is_active = False
        account.save()
        messages.success(request, 'Account deleted successfully!')
        return redirect('dashboard:accounts_list')
    
    return render(request, 'dashboard/account_confirm_delete.html', {'account': account})


@login_required
def transactions_list(request):
    """View to list all transactions"""
    transactions = Transaction.objects.filter(user=request.user).order_by('-date', '-created_at')
    
    # Filter options
    account_filter = request.GET.get('account')
    category_filter = request.GET.get('category')
    transaction_type_filter = request.GET.get('transaction_type')
    
    if account_filter:
        transactions = transactions.filter(account_id=account_filter)
    if category_filter:
        transactions = transactions.filter(category=category_filter)
    if transaction_type_filter:
        transactions = transactions.filter(transaction_type=transaction_type_filter)
    
    accounts = Account.objects.filter(user=request.user, is_active=True)
    
    # Calculate summary statistics
    total_income = transactions.filter(transaction_type='income').aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = transactions.filter(transaction_type='expense').aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'transactions': transactions,
        'accounts': accounts,
        'categories': Transaction.CATEGORIES,
        'transaction_types': Transaction.TRANSACTION_TYPES,
        'account_filter': account_filter,
        'category_filter': category_filter,
        'transaction_type_filter': transaction_type_filter,
        'total_income': total_income,
        'total_expenses': total_expenses,
    }
    return render(request, 'dashboard/transactions_list.html', context)


@login_required
def transaction_create(request):
    """View to create a new transaction"""
    if request.method == 'POST':
        form = TransactionForm(request.POST, user=request.user)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user
            transaction.save()
            messages.success(request, 'Transaction added successfully!')
            return redirect('dashboard:transactions_list')
    else:
        form = TransactionForm(user=request.user)
    
    return render(request, 'dashboard/transaction_form.html', {
        'form': form, 
        'title': 'Add New Transaction'
    })


@login_required
def transaction_edit(request, transaction_id):
    """View to edit an existing transaction"""
    transaction = get_object_or_404(Transaction, id=transaction_id, user=request.user)
    
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Transaction updated successfully!')
            return redirect('dashboard:transactions_list')
    else:
        form = TransactionForm(instance=transaction, user=request.user)
    
    return render(request, 'dashboard/transaction_form.html', {
        'form': form, 
        'title': 'Edit Transaction',
        'transaction': transaction
    })


@login_required
def transaction_delete(request, transaction_id):
    """View to delete a transaction"""
    transaction = get_object_or_404(Transaction, id=transaction_id, user=request.user)
    
    if request.method == 'POST':
        transaction.delete()
        messages.success(request, 'Transaction deleted successfully!')
        return redirect('dashboard:transactions_list')
    
    return render(request, 'dashboard/transaction_confirm_delete.html', {'transaction': transaction})


@login_required
def account_entries(request):
    """View to manage account entries for current month"""
    user = request.user
    current_month = timezone.now().month
    current_year = timezone.now().year
    
    # Get selected month/year from request
    if request.method == 'POST':
        selected_month = int(request.POST.get('month', request.GET.get('month', current_month)))
        selected_year = int(request.POST.get('year', request.GET.get('year', current_year)))
    else:
        selected_month = int(request.GET.get('month', current_month))
        selected_year = int(request.GET.get('year', current_year))
    
    # Get sorting parameters
    sort_by = request.GET.get('sort', 'name')
    order = request.GET.get('order', 'asc')
    
    # Validate sort field
    valid_sort_fields = ['name', 'account_type', 'classification', 'asset_type', 'institution', 'balance']
    if sort_by not in valid_sort_fields:
        sort_by = 'name'
    
    # Get all user accounts with sorting
    accounts = Account.objects.filter(user=user, is_active=True)
    
    # Apply sorting
    if sort_by == 'balance':
        # For balance sorting, we need to annotate with the latest balance
        accounts = accounts.annotate(
            latest_balance=models.Subquery(
                AccountEntry.objects.filter(
                    account=models.OuterRef('pk')
                ).order_by('-year', '-month').values('balance')[:1]
            )
        )
        if order == 'desc':
            accounts = accounts.order_by('-latest_balance')
        else:
            accounts = accounts.order_by('latest_balance')
    else:
        if order == 'desc':
            accounts = accounts.order_by(f'-{sort_by}')
        else:
            accounts = accounts.order_by(sort_by)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'clear_data':
            # Handle clear data action
            cleared_count = 0
            for account in accounts:
                entry, created = AccountEntry.objects.get_or_create(
                    account=account,
                    month=selected_month,
                    year=selected_year,
                    defaults={'balance': 0.00, 'notes': ''}
                )
                entry.balance = 0.00
                entry.notes = ''
                entry.save()
                cleared_count += 1
            
            if cleared_count > 0:
                messages.success(request, f'Successfully cleared {cleared_count} account entries for {selected_month}/{selected_year}')
            return redirect(f'{reverse("dashboard:account_entries")}?month={selected_month}&year={selected_year}&sort={sort_by}&order={order}')
        
        else:
            # Handle form submission for multiple accounts
            success_count = 0
            for account in accounts:
                balance_key = f'balance_{account.id}'
                notes_key = f'notes_{account.id}'
                
                if balance_key in request.POST:
                    balance = request.POST.get(balance_key)
                    notes = request.POST.get(notes_key, '')
                    
                    if balance:  # Only save if balance is provided
                        entry, created = AccountEntry.objects.get_or_create(
                            account=account,
                            month=selected_month,
                            year=selected_year,
                            defaults={'balance': balance, 'notes': notes}
                        )
                        if not created:
                            entry.balance = balance
                            entry.notes = notes
                            entry.save()
                        success_count += 1
            
            if success_count > 0:
                messages.success(request, f'Successfully updated {success_count} account entries for {selected_month}/{selected_year}')
            return redirect(f'{reverse("dashboard:account_entries")}?month={selected_month}&year={selected_year}&sort={sort_by}&order={order}')
    
    # Get existing entries for the selected month/year
    entries = {}
    for account in accounts:
        # Get the actual entry for the selected month/year
        entry = AccountEntry.objects.filter(
            account=account,
            month=selected_month,
            year=selected_year
        ).first()
        
        if not entry:
            # If no entry exists for this month, create a placeholder with 0 balance
            entry = AccountEntry(
                account=account,
                month=selected_month,
                year=selected_year,
                balance=0.00,
                notes=''
            )
        
        entries[account.id] = entry
    
    # Group accounts by classification
    accounts_by_classification = {}
    classification_totals = {}
    
    for account in accounts:
        classification = account.get_classification_display()
        if classification not in accounts_by_classification:
            accounts_by_classification[classification] = []
            classification_totals[classification] = 0
        accounts_by_classification[classification].append(account)
        
        # Calculate total for this classification using the entry for this month
        entry = entries.get(account.id)
        if entry and entry.balance:
            classification_totals[classification] += entry.balance
    
    context = {
        'accounts': accounts,
        'accounts_by_classification': accounts_by_classification,
        'classification_totals': classification_totals,
        'entries': entries,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'current_month': current_month,
        'current_year': current_year,
        'sort_by': sort_by,
        'order': order,
    }
    
    return render(request, 'dashboard/account_entries.html', context)


@login_required
def settings(request):
    """View for account settings"""
    user = request.user
    
    if request.method == 'POST':
        # Handle form submission for updating user profile
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        
        # Basic validation
        if not email:
            messages.error(request, 'Email is required.')
            return redirect('dashboard:settings')
        
        # Check if email is already taken by another user
        if User.objects.filter(email=email).exclude(id=user.id).exists():
            messages.error(request, 'This email is already in use by another account.')
            return redirect('dashboard:settings')
        
        # Update user profile
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        user.save()
        
        messages.success(request, 'Profile updated successfully!')
        return redirect('dashboard:settings')
    
    # Get account statistics
    accounts_count = Account.objects.filter(user=user, is_active=True).count()
    transactions_count = Transaction.objects.filter(account__user=user).count()
    entries_count = AccountEntry.objects.filter(account__user=user).count()
    
    context = {
        'user': user,
        'accounts_count': accounts_count,
        'transactions_count': transactions_count,
        'entries_count': entries_count,
    }
    
    return render(request, 'dashboard/settings.html', context)


@login_required
def data_management(request):
    """View for data management options"""
    user = request.user
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'delete_accounts':
            # Delete all user's accounts and related data
            accounts = Account.objects.filter(user=user)
            accounts_count = accounts.count()
            
            # Delete related entries and transactions
            AccountEntry.objects.filter(account__user=user).delete()
            Transaction.objects.filter(account__user=user).delete()
            accounts.delete()
            
            messages.success(request, f'Successfully deleted {accounts_count} accounts and all related data.')
            return redirect('dashboard:dashboard')
            
        elif action == 'delete_entries':
            # Delete all account entries
            entries = AccountEntry.objects.filter(account__user=user)
            entries_count = entries.count()
            entries.delete()
            
            messages.success(request, f'Successfully deleted {entries_count} account entries.')
            return redirect('dashboard:data_management')
            
        elif action == 'import_accounts':
            # Handle accounts import
            if 'csv_file' in request.FILES:
                csv_file = request.FILES['csv_file']
                if csv_file.name.endswith('.csv'):
                    try:
                        # Decode the file content
                        decoded_file = csv_file.read().decode('utf-8').splitlines()
                        reader = csv.DictReader(decoded_file)
                        
                        imported_count = 0
                        errors = []
                        
                        for row in reader:
                            try:
                                # Create account from CSV row
                                account = Account(
                                    user=user,
                                    name=row.get('name', '').strip(),
                                    account_type=row.get('account_type', 'checking'),
                                    classification=row.get('classification', 'taxable'),
                                    asset_type=row.get('asset_type', 'cash'),
                                    currency=row.get('currency', 'USD'),
                                    institution=row.get('institution', '').strip(),
                                    account_number=row.get('account_number', '').strip(),
                                )
                                account.save()
                                imported_count += 1
                                
                            except Exception as e:
                                errors.append(f"Row {reader.line_num}: {str(e)}")
                        
                        if imported_count > 0:
                            messages.success(request, f'Successfully imported {imported_count} accounts.')
                        if errors:
                            messages.warning(request, f'Import completed with {len(errors)} errors. Check the data format.')
                            
                    except Exception as e:
                        messages.error(request, f'Error importing accounts: {str(e)}')
                else:
                    messages.error(request, 'Please upload a valid CSV file.')
            else:
                messages.error(request, 'Please select a CSV file to import.')
                
        elif action == 'import_entries':
            # Handle account entries import
            print(f"DEBUG: Starting import_entries action")
            if 'csv_file' in request.FILES:
                csv_file = request.FILES['csv_file']
                print(f"DEBUG: CSV file received: {csv_file.name}, size: {csv_file.size}")
                if csv_file.name.endswith('.csv'):
                    try:
                        # Decode the file content
                        decoded_file = csv_file.read().decode('utf-8').splitlines()
                        print(f"DEBUG: Decoded {len(decoded_file)} lines from CSV")
                        reader = csv.DictReader(decoded_file)
                        print(f"DEBUG: CSV headers: {reader.fieldnames}")
                        
                        # Validate CSV headers
                        required_headers = ['account_name', 'month', 'year', 'balance']
                        if not reader.fieldnames:
                            messages.error(request, 'CSV file appears to be empty or malformed.')
                            return render(request, 'dashboard/data_management.html', context)
                            
                        missing_headers = [h for h in required_headers if h not in reader.fieldnames]
                        if missing_headers:
                            messages.error(request, f'CSV is missing required headers: {", ".join(missing_headers)}. Found headers: {", ".join(reader.fieldnames)}')
                            return render(request, 'dashboard/data_management.html', context)
                        
                        imported_count = 0
                        errors = []
                        
                        for row_num, row in enumerate(reader, start=2):  # Start at 2 because row 1 is header
                            print(f"DEBUG: Processing row {row_num}: {row}")
                            
                            # Skip empty rows
                            if not row or all(not value.strip() for value in row.values()):
                                print(f"DEBUG: Skipping empty row {row_num}")
                                continue
                                
                            try:
                                # Find the account by name
                                account_name = row.get('account_name', '').strip()
                                print(f"DEBUG: Row {row_num} - account_name: '{account_name}'")
                                if not account_name:
                                    errors.append(f"Row {row_num}: Missing account name")
                                    continue
                                    
                                account = Account.objects.filter(user=user, name=account_name).first()
                                print(f"DEBUG: Row {row_num} - account found: {account}")
                                
                                if account:
                                    # Parse month and year
                                    month_raw = row.get('month', 1)
                                    month = parse_month(month_raw)
                                    print(f"DEBUG: Row {row_num} - month_raw: '{month_raw}', parsed: {month}")
                                    
                                    year_str = row.get('year', '2024')
                                    print(f"DEBUG: Row {row_num} - year_str: '{year_str}'")
                                    try:
                                        year = int(year_str)
                                        if year < 1900 or year > 2100:
                                            errors.append(f"Row {row_num}: Invalid year {year_str}")
                                            continue
                                    except ValueError:
                                        errors.append(f"Row {row_num}: Invalid year {year_str}")
                                        continue
                                    
                                    balance_raw = row.get('balance', 0)
                                    balance = parse_balance(balance_raw)
                                    print(f"DEBUG: Row {row_num} - balance_raw: '{balance_raw}', parsed: {balance}")
                                    
                                    notes = row.get('notes', '').strip()
                                    print(f"DEBUG: Row {row_num} - notes: '{notes}'")
                                    
                                    # Create or update entry from CSV row
                                    entry, created = AccountEntry.objects.get_or_create(
                                        account=account,
                                        month=month,
                                        year=year,
                                        defaults={
                                            'balance': balance,
                                            'notes': notes
                                        }
                                    )
                                    
                                    if not created:
                                        # Update existing entry
                                        entry.balance = balance
                                        entry.notes = notes
                                        entry.save()
                                    
                                    imported_count += 1
                                else:
                                    errors.append(f"Row {row_num}: Account '{account_name}' not found")
                                    
                            except Exception as e:
                                errors.append(f"Row {row_num}: {str(e)}")
                                print(f"Import error on row {row_num}: {str(e)}")
                                print(f"Row data: {row}")
                        
                        print(f"DEBUG: Import completed - imported: {imported_count}, errors: {len(errors)}")
                        if errors:
                            print(f"DEBUG: Error details: {errors}")
                            
                        if imported_count > 0:
                            messages.success(request, f'Successfully imported {imported_count} account entries.')
                        if errors:
                            messages.warning(request, f'Import completed with {len(errors)} errors. Check the data format.')
                            
                    except Exception as e:
                        messages.error(request, f'Error importing entries: {str(e)}')
                else:
                    messages.error(request, 'Please upload a valid CSV file.')
            else:
                messages.error(request, 'Please select a CSV file to import.')
    
    # Get data statistics
    accounts_count = Account.objects.filter(user=user, is_active=True).count()
    transactions_count = Transaction.objects.filter(account__user=user).count()
    entries_count = AccountEntry.objects.filter(account__user=user).count()
    
    context = {
        'accounts_count': accounts_count,
        'transactions_count': transactions_count,
        'entries_count': entries_count,
    }
    
    return render(request, 'dashboard/data_management.html', context)


@login_required
def export_csv(request, data_type):
    """Export data as CSV"""
    user = request.user
    
    if data_type == 'accounts':
        data = Account.objects.filter(user=user, is_active=True)
        filename = f'accounts_{user.username}_{timezone.now().strftime("%Y%m%d")}.csv'
        fieldnames = ['name', 'account_type', 'classification', 'asset_type', 'currency', 'institution', 'account_number', 'created_at']
        
    elif data_type == 'transactions':
        data = Transaction.objects.filter(account__user=user)
        filename = f'transactions_{user.username}_{timezone.now().strftime("%Y%m%d")}.csv'
        fieldnames = ['account__name', 'amount', 'transaction_type', 'category', 'description', 'date', 'created_at']
        
    elif data_type == 'entries':
        data = AccountEntry.objects.filter(account__user=user)
        filename = f'account_entries_{user.username}_{timezone.now().strftime("%Y%m%d")}.csv'
        fieldnames = ['account__name', 'month', 'year', 'balance', 'notes', 'created_at']
    
    else:
        return HttpResponse('Invalid data type', status=400)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.DictWriter(response, fieldnames=fieldnames)
    writer.writeheader()
    
    for item in data:
        row = {}
        for field in fieldnames:
            if '__' in field:
                # Handle related fields
                parts = field.split('__')
                value = item
                for part in parts:
                    value = getattr(value, part)
                row[field] = str(value) if value else ''
            else:
                value = getattr(item, field)
                # Special handling for month field in entries export
                if data_type == 'entries' and field == 'month':
                    row[field] = get_month_name(value)
                else:
                    row[field] = str(value) if value else ''
        writer.writerow(row)
    
    return response


@login_required
def export_excel(request, data_type):
    """Export data as Excel"""
    user = request.user
    
    if data_type == 'accounts':
        data = Account.objects.filter(user=user, is_active=True)
        filename = f'accounts_{user.username}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        headers = ['Name', 'Account Type', 'Classification', 'Asset Type', 'Currency', 'Institution', 'Account Number', 'Created At']
        
    elif data_type == 'transactions':
        data = Transaction.objects.filter(account__user=user)
        filename = f'transactions_{user.username}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        headers = ['Account', 'Amount', 'Type', 'Category', 'Description', 'Date', 'Created At']
        
    elif data_type == 'entries':
        data = AccountEntry.objects.filter(account__user=user)
        filename = f'account_entries_{user.username}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        headers = ['Account', 'Month', 'Year', 'Balance', 'Notes', 'Created At']
    
    else:
        return HttpResponse('Invalid data type', status=400)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = data_type.title()
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, item in enumerate(data, 2):
        if data_type == 'accounts':
            ws.cell(row=row, column=1, value=item.name)
            ws.cell(row=row, column=2, value=item.get_account_type_display())
            ws.cell(row=row, column=3, value=item.get_classification_display())
            ws.cell(row=row, column=4, value=item.get_asset_type_display())
            ws.cell(row=row, column=5, value=item.currency)
            ws.cell(row=row, column=6, value=item.institution or '')
            ws.cell(row=row, column=7, value=item.account_number or '')
            ws.cell(row=row, column=8, value=item.created_at.strftime('%Y-%m-%d %H:%M'))
            
        elif data_type == 'transactions':
            ws.cell(row=row, column=1, value=item.account.name)
            ws.cell(row=row, column=2, value=float(item.amount))
            ws.cell(row=row, column=3, value=item.get_transaction_type_display())
            ws.cell(row=row, column=4, value=item.get_category_display())
            ws.cell(row=row, column=5, value=item.description)
            ws.cell(row=row, column=6, value=item.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=7, value=item.created_at.strftime('%Y-%m-%d %H:%M'))
            
        elif data_type == 'entries':
            ws.cell(row=row, column=1, value=item.account.name)
            ws.cell(row=row, column=2, value=item.month)
            ws.cell(row=row, column=3, value=item.year)
            ws.cell(row=row, column=4, value=float(item.balance))
            ws.cell(row=row, column=5, value=item.notes or '')
            ws.cell(row=row, column=6, value=item.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_pdf(request, data_type):
    """Export data as PDF"""
    user = request.user
    
    if data_type == 'accounts':
        data = Account.objects.filter(user=user, is_active=True)
        filename = f'accounts_{user.username}_{timezone.now().strftime("%Y%m%d")}.pdf'
        title = 'Accounts Report'
        headers = ['Name', 'Type', 'Classification', 'Asset Type', 'Institution']
        
    elif data_type == 'transactions':
        data = Transaction.objects.filter(account__user=user)
        filename = f'transactions_{user.username}_{timezone.now().strftime("%Y%m%d")}.pdf'
        title = 'Transactions Report'
        headers = ['Account', 'Amount', 'Type', 'Category', 'Description', 'Date']
        
    elif data_type == 'entries':
        data = AccountEntry.objects.filter(account__user=user)
        filename = f'account_entries_{user.username}_{timezone.now().strftime("%Y%m%d")}.pdf'
        title = 'Account Entries Report'
        headers = ['Account', 'Month/Year', 'Balance', 'Notes']
    
    else:
        return HttpResponse('Invalid data type', status=400)
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Add title
    elements.append(Paragraph(f"NetWorth Tracker - {title}", title_style))
    elements.append(Spacer(1, 20))
    
    # Prepare table data
    table_data = [headers]
    
    for item in data:
        if data_type == 'accounts':
            row = [
                item.name,
                item.get_account_type_display(),
                item.get_classification_display(),
                item.get_asset_type_display(),
                item.institution or 'N/A'
            ]
        elif data_type == 'transactions':
            row = [
                item.account.name,
                f"${item.amount:,.2f}",
                item.get_transaction_type_display(),
                item.get_category_display(),
                item.description,
                item.date.strftime('%Y-%m-%d')
            ]
        elif data_type == 'entries':
            row = [
                item.account.name,
                f"{item.month}/{item.year}",
                f"${item.balance:,.2f}",
                item.notes or 'N/A'
            ]
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Add summary
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=10
    )
    elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", summary_style))
    elements.append(Paragraph(f"Total records: {len(data)}", summary_style))
    
    doc.build(elements)
    return response


@login_required
def support(request):
    """View for support and donation page"""
    return render(request, 'dashboard/support.html')


def terms_of_service(request):
    """View for Terms of Service page"""
    return render(request, 'dashboard/terms_of_service.html')


def privacy_policy(request):
    """View for Privacy Policy page"""
    return render(request, 'dashboard/privacy_policy.html')



